from __future__ import annotations

from calendar import monthrange
from datetime import datetime, time
from pathlib import Path
from typing import Dict, Iterable, Optional
import logging

from openpyxl import load_workbook

from models import AttendanceRecord
from utils import extract_time_text


class ExcelExportError(RuntimeError):
    pass


def _excel_time(value: Optional[str]) -> Optional[time]:
    normalized = extract_time_text(value or "")
    if not normalized:
        return None
    hours, minutes = (int(part) for part in normalized.split(":"))
    return time(hour=hours, minute=minutes)


def export_excel(
    template_path: str | Path,
    output_path: str | Path,
    name: str,
    month: int,
    records: Iterable[AttendanceRecord],
    mapping: Dict,
) -> None:
    try:
        workbook = load_workbook(template_path)
        sheet_name = mapping.get("sheet_name")
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

        name_cell = mapping.get("name_cell")
        if name_cell:
            prefix = str(mapping.get("name_prefix", ""))
            worksheet[name_cell] = f"{prefix}{name}".rstrip(".")

        month_cell = mapping.get("month_cell")
        if month_cell:
            worksheet[month_cell] = int(month)

        start_row = int(mapping.get("start_row", 3))
        start_col = mapping.get("start_col", "B")
        end_col = mapping.get("end_col", "C")
        second_start_col = mapping.get("second_start_col", "D")
        second_end_col = mapping.get("second_end_col", "E")
        hours_col = mapping.get("hours_col", "F")
        unit_col = mapping.get("unit_col", "G")
        date_col = mapping.get("date_col", "A")

        today = datetime.now()
        year = today.year
        valid_days = monthrange(year, int(month))[1]
        by_day = {record.day: record for record in records}

        for day in range(1, 32):
            row = start_row + day - 1
            if date_col:
                worksheet[f"{date_col}{row}"] = day if day <= valid_days else None
            record = by_day.get(day)
            start_value = _excel_time(record.start_time if record else None)
            end_value = _excel_time(record.end_time if record else None)
            worksheet[f"{start_col}{row}"] = start_value
            worksheet[f"{end_col}{row}"] = end_value
            worksheet[f"{second_start_col}{row}"] = None
            worksheet[f"{second_end_col}{row}"] = None
            for col in (start_col, end_col, second_start_col, second_end_col):
                worksheet[f"{col}{row}"].number_format = "h:mm"
            worksheet[f"{hours_col}{row}"] = f"=({end_col}{row}-{start_col}{row})+({second_end_col}{row}-{second_start_col}{row})"
            worksheet[f"{unit_col}{row}"] = f"=HOUR({hours_col}{row})+MINUTE({hours_col}{row})/60"

        total_cell = mapping.get("total_cell")
        if total_cell:
            last_row = start_row + 30
            worksheet[total_cell] = f"=SUM({unit_col}{start_row}:{unit_col}{last_row})"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        logging.info("exported Excel: %s", output_path)
    except PermissionError as exc:
        raise ExcelExportError("Could not write the Excel file. Please close the output workbook and try again.") from exc
    except Exception as exc:
        logging.exception("Excel export failed")
        raise ExcelExportError(f"Excel export failed: {exc}") from exc
